from openpyxl import load_workbook

class IngestRequest:
    def __init__(self, filepath):
        self.wb = load_workbook(filepath)

    def print_sheet_names(self):
       print(self.wb.sheetnames)